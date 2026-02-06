import os
import json
from datetime import datetime

import gspread
from google.oauth2.service_account import Credentials
from openpyxl import Workbook
from openpyxl.styles import Font

from bot.tgbot.databases.pay_db import get_user_by_user_id
from bot.tgbot.databases.database import get_db_connection
from config import (
    BASE_DIR,
    MAIN_DB_PATH,
    ADVERT_TOKENS_DB_PATH,
    ADVERT_POSITIONS_FILE,
    DB_TYPE,
    logger_bot,
)


# Настройки Google Sheets
GOOGLE_CREDENTIALS_FILE = os.path.join(
    BASE_DIR,
    "bot",
    "bot_prodazhi",
    "service_account.json",
)
SPREADSHEET_URL = "https://docs.google.com/spreadsheets/d/1k59CdZ0vdCcJnZMdXczKoQe2ExD26_wIHSm7lWxPRn0/edit?gid=473182526"

scope = [
    "https://spreadsheets.google.com/feeds",
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive.file",
    "https://www.googleapis.com/auth/drive",
]


def getUnpaids():
    """Получает список неоплативших пользователей"""
    db = get_db_connection(MAIN_DB_PATH, schema="main")
    usernames = db.fetchall("SELECT full_name FROM users WHERE pay_status::int = 0")
    return usernames


def create_excel():
    # Получаем текущий путь к скрипту
    script_dir = os.path.dirname(os.path.abspath(__file__))
    # Создаем новую книгу Excel
    wb = Workbook()
    # Выбираем активный лист
    sheet = wb.active
    # Устанавливаем заголовок для столбца
    sheet.merge_cells("A1:B1")
    sheet["A1"] = "Не оплаченные:"  # Пишем заголовок в объединенные ячейки
    sheet["A1"].font = Font(size=15)

    # Добавляем данные
    unpaids = getUnpaids()
    counter = 0

    # Записываем данные из списка unpaids в столбец A таблицы Excel
    for idx, unpaid in enumerate(
        unpaids, start=2
    ):  # Начинаем с 2, чтобы не перезаписать заголовок
        sheet[f"A{idx}"] = unpaid["full_name"] # unpaid[0]  # Юзернейм
        counter += 1

    sheet["D1"] = "Всего:"  # Пишем заголовок в объединенные ячейки
    sheet["D1"].font = Font(size=19)
    sheet["E1"] = f"{counter}"  # Пишем заголовок в объединенные ячейки
    sheet["E1"].font = Font(size=20)

    # Сохраняем книгу возле скрипта
    filename = "dataunpaids.xlsx"
    filepath = os.path.join(script_dir, filename)
    wb.save(filepath)
    logger_bot.info(f"Excel-таблица создана и сохранена как '{filepath}'")


def getpaids():
    """Получает список оплативших пользователей"""
    db = get_db_connection(MAIN_DB_PATH, schema="main")
    if DB_TYPE == "postgres":
        query = """
        SELECT
            full_name,
            full_name_payments,
            TO_CHAR(TO_TIMESTAMP(last_pay), 'DD-MM-YYYY HH24:MI') as last_pay,
            TO_CHAR(TO_TIMESTAMP(end_pay), 'DD-MM-YYYY HH24:MI') as end_pay
        FROM
            users
        WHERE
            pay_status::int = 1
        ORDER BY
            last_pay ASC
        """
    else:
        query = """
        SELECT
            full_name,
            fullName,
            strftime('%%d-%%m-%%Y %%H:%%M', datetime(last_pay, 'unixepoch')) as last_pay,
            strftime('%%d-%%m-%%Y %%H:%%M', datetime(end_pay, 'unixepoch')) as end_pay
        FROM
            users
        WHERE
            pay_status = 1
        ORDER BY
            last_pay ASC
        """
    result = db.fetchall(query)
    return result


def create_excel1():
    # Получаем текущий путь к скрипту
    script_dir = os.path.dirname(os.path.abspath(__file__))
    # Создаем новую книгу Excel
    wb = Workbook()
    # Выбираем активный лист
    sheet = wb.active
    # Устанавливаем заголовок для столбца
    sheet.merge_cells("A1:B1")
    sheet["A1"] = "оплаченные:"  # Пишем заголовок в объединенные ячейки
    sheet["A1"].font = Font(size=15)

    # Добавляем данные
    paids = (
        getpaids()
    )  # Предполагаем, что функция возвращает данные в формате (name, last_pay, end_pay)
    counter = 0

    # Добавляем заголовки столбцов
    sheet["A2"] = "Имя"
    sheet["B2"] = "username"
    sheet["C2"] = "Когда оплатил"
    sheet["D2"] = "До какого оплатил"

    # Записываем данные
    for idx, paid in enumerate(
        paids, start=3
    ):  # Начинаем с 2 строки (после заголовков)
        # sheet[f"A{idx}"] = unpaid[0]
        # sheet[f"B{idx}"] = unpaid[1]  # Имя
        # sheet[f"C{idx}"] = unpaid[2]  # Дата оплаты
        # sheet[f"D{idx}"] = unpaid[3]  # Дата окончания
        sheet[f"A{idx}"] = paid.get("full_name", "")
        sheet[f"B{idx}"] = paid.get("full_name_payments", "")
        sheet[f"C{idx}"] = paid.get("last_pay", "")
        sheet[f"D{idx}"] = paid.get("end_pay", "")
        counter += 1

    sheet["E1"] = "Всего:"  # Пишем заголовок в объединенные ячейки
    sheet["E1"].font = Font(size=19)
    sheet["F1"] = f"{counter}"  # Пишем заголовок в объединенные ячейки
    sheet["F1"].font = Font(size=20)

    # Сохраняем книгу возле скрипта
    filename = "datapaids.xlsx"
    filepath = os.path.join(script_dir, filename)
    wb.save(filepath)
    logger_bot.info(f"Excel-таблица создана и сохранена как '{filepath}'")


def get_lawyer_requests():
    """Получает запросы юриста за последние 2 месяца"""
    db = get_db_connection(MAIN_DB_PATH, schema="main")
    if DB_TYPE == "postgres":
        query = """
        SELECT 
            TO_CHAR(request_date, 'DD-MM-YYYY HH24:MI') as formatted_date,
            request_text,
            user_full_name,
            user_username
        FROM 
            requests 
        WHERE 
            request_type = 'lawyer'
            AND request_date >= NOW() - INTERVAL '2 months'
        ORDER BY 
            request_date DESC
        """
    else:
        query = """
        SELECT 
            strftime('%%d-%%m-%%Y %%H:%%M', request_date) as formatted_date,
            request_text,
            user_full_name,
            user_username
        FROM 
            requests 
        WHERE 
            request_type = 'lawyer'
            AND request_date >= datetime('now', '-2 months')
        ORDER BY 
            request_date DESC
        """
    result = db.fetchall(query)
    return result


def create_excel_lawyer():
    # Получаем текущий путь к скрипту
    script_dir = os.path.dirname(os.path.abspath(__file__))
    # Создаем новую книгу Excel
    wb = Workbook()
    # Выбираем активный лист
    sheet = wb.active
    # Устанавливаем заголовок для столбца

    # Добавляем данные
    unpaids = (
        get_lawyer_requests()
    )  # Предполагаем, что функция возвращает данные в формате (name, last_pay, end_pay)
    counter = 0

    # Добавляем заголовки столбцов
    sheet["A1"] = "дата"
    sheet["B1"] = "текст запроса"
    sheet["C1"] = "имя"
    sheet["D1"] = "username"

    # Записываем данные
    for idx, unpaid in enumerate(
        unpaids, start=2
    ):  # Начинаем с 2 строки (после заголовков)
        sheet[f"A{idx}"] = unpaid.get("formatted_date", "")
        sheet[f"B{idx}"] = unpaid.get("request_text", "")
        sheet[f"C{idx}"] = unpaid.get("user_full_name", "")
        sheet[f"D{idx}"] = unpaid.get("user_username", "")
        counter += 1

    # Сохраняем книгу возле скрипта
    filename = "lawyer.xlsx"
    filepath = os.path.join(script_dir, filename)
    wb.save(filepath)
    logger_bot.info(f"Excel-таблица создана и сохранена как '{filepath}'")


def load_positions():
    """Загружает конфигурацию позиций из JSON."""
    try:
        with open(ADVERT_POSITIONS_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
            return data.get("positions", [])
    except Exception as e:
        logger_bot.error(f"Ошибка при загрузке JSON с позициями: {e}")
        return []


def get_advert_requests():
    """Получает запросы на рекламу за последние 2 месяца"""
    db = get_db_connection(MAIN_DB_PATH, schema="main")
    if DB_TYPE == "postgres":
        query = """
        SELECT
            TO_CHAR(request_date, 'DD-MM-YYYY HH24:MI') as formatted_date,
            request_text,
            user_full_name,
            user_username
        FROM
            requests
        WHERE
            request_type = 'advert'
            AND request_date >= NOW() - INTERVAL '2 months'
        ORDER BY
            request_date DESC
        """
    else:
        query = """
        SELECT
            strftime('%%d-%%m-%%Y %%H:%%M', request_date) as formatted_date,
            request_text,
            user_full_name,
            user_username
        FROM
            requests
        WHERE
            request_type = 'advert'
            AND request_date >= datetime('now', '-2 months')
        ORDER BY
            request_date DESC
        """
    result = db.fetchall(query)
    return result


def get_advert_requests_new():
    """Получаем оплаченные объявления из таблицы tokens"""
    db = get_db_connection(ADVERT_TOKENS_DB_PATH, schema="advert")
    result = db.fetchall(
        """
        SELECT
            created_at,
            user_id,
            data_json
        FROM tokens
        WHERE payment_status::int = 1
        ORDER BY created_at DESC
        """
    )
    return result


def create_google_sheets():
    creds = Credentials.from_service_account_file(
        GOOGLE_CREDENTIALS_FILE,
        scopes=scope,
    )
    gs_client = gspread.authorize(creds)
    spreadsheet = gs_client.open_by_url(SPREADSHEET_URL)

    # создадим (или перезапишем) лист под отчёт
    try:
        sheet_advert = spreadsheet.worksheet("Отчет по рекламе")
        sheet_advert.clear()  # очищаем старые данные
    except gspread.exceptions.WorksheetNotFound:
        sheet_advert = spreadsheet.add_worksheet(
            title="Отчет по рекламе", rows="1000", cols="20"
        )
    return sheet_advert


def create_excel_advert():
    # Получаем текущий путь к скрипту
    script_dir = os.path.dirname(os.path.abspath(__file__))
    # Создаем новую книгу Excel
    wb = Workbook()
    # Выбираем активный лист
    sheet = wb.active
    # Устанавливаем заголовок для столбца

    # Добавляем данные
    unpaids = (
        get_advert_requests()
    )  # Предполагаем, что функция возвращает данные в формате (name, last_pay, end_pay)
    counter = 0

    # Добавляем заголовки столбцов
    sheet["A1"] = "дата"
    sheet["B1"] = "текст запроса"
    sheet["C1"] = "имя"
    sheet["D1"] = "username"

    # Записываем данные
    for idx, unpaid in enumerate(
        unpaids, start=2
    ):  # Начинаем с 2 строки (после заголовков)
        sheet[f"A{idx}"] = unpaid.get("formatted_date", "")
        sheet[f"B{idx}"] = unpaid.get("request_text", "")
        sheet[f"C{idx}"] = unpaid.get("user_full_name", "")
        sheet[f"D{idx}"] = unpaid.get("user_username", "")
        counter += 1

    # Сохраняем книгу возле скрипта
    filename = "advert.xlsx"
    filepath = os.path.join(script_dir, filename)
    wb.save(filepath)
    print(f"Excel-таблица создана и сохранена как '{filepath}'")


def create_excel_advert_new():
    # sheet_advert = create_google_sheets()

    """Создаёт Excel-файл с отчётом по объявлениям"""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Advert Report"

    # 🔹 Загружаем конфигурацию позиций
    positions = load_positions()

    # 🔹 Определяем порядок площадок вручную
    ordered_position_keys = [p["key"] for p in positions]

    # 🔹 Формируем заголовки в нужном порядке
    start_headers = ["Дата платежа", "Имя", "Username"]
    position_headers = [p["name"] for p in positions]
    end_header = [
        "Размещений Авито",
        "Размещений ЦИАН",
        "Всего размещений",
        "Оплачено Авито",
        "Оплачено ЦИАН",
        "Всего оплачено",
    ]
    full_headers = start_headers + position_headers + end_header

    sheet.append(full_headers)
    # sheet_advert.append_row(full_headers)

    # 🔹 Получаем данные из БД
    records = get_advert_requests_new()

    # TODO Тут можно вручную изменить дату формирования отчёт
    target_date = datetime(2026, 1, 19, 19, 0, 0)
    #target_date = datetime.today()
    for row in records:
        created_at, user_id, data_json_str = row
        created_at_dt = datetime.strptime(created_at, "%Y-%m-%d %H:%M:%S")
        if created_at_dt < target_date:
            continue

        user_data = get_user_by_user_id(user_id)
        username = user_data.get("fullName", "")

        outer = json.loads(data_json_str)
        inner = json.loads(outer.get("data_json", "{}"))
        full_name = inner.get("full_name", "")

        if not full_name:
            full_name = user_data.get("full_name", "")

        # Безопасно парсим JSON
        try:
            outer = json.loads(data_json_str)
            data = json.loads(outer) if isinstance(outer, str) else outer
        except Exception:
            data = {}

        json_data_str = data.get("data_json", "{}")
        try:
            json_data = json.loads(json_data_str)
        except Exception:
            json_data = {}

        # 🔹 Подсчёты
        total_ads = 0
        avito_ads = 0
        cian_ads = 0
        paid_avito = 0.0
        paid_cian = 0.0

        # Словарь цен для быстрого доступа
        price_map = {p["key"]: float(p.get("price", 0)) for p in positions}

        # 🔹 Данные по позициям в нужном порядке
        position_values = []
        for key in ordered_position_keys:
            value_str = str(json_data.get(key, "0"))
            try:
                count = int(value_str)
            except ValueError:
                count = 0

            position_values.append(count)
            total_ads += count

            # Распределяем по площадкам
            price = price_map.get(key, 0)
            if "avito" in key.lower():
                avito_ads += count
                paid_avito += count * price
            elif "cian" in key.lower():
                cian_ads += count
                paid_cian += count * price

        total_price = json_data.get("total_price", "0")
        if total_price == "0":
            continue

        # 🔹 Формируем строку данных в нужном порядке
        row_data = (
            [
                created_at,
                full_name,
                username,
            ]
            + position_values
            + [
                avito_ads,
                cian_ads,
                total_ads,
                paid_avito,
                paid_cian,
                total_price,
            ]
        )

        sheet.append(row_data)
        # sheet_advert.append_row(row_data)
    # logger_bot.info("✅ Отчет по запросам на рекламу сохранён в Google таблицу")

    # 🔹 Форматируем ширину колонок
    for col in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        sheet.column_dimensions[col[0].column_letter].width = max_length + 2

    # 🔹 Сохраняем файл
    filepath = os.path.join(script_dir, "advert_report.xlsx")
    wb.save(filepath)
    logger_bot.info(f"✅ Отчет по запросам на рекламу сохранён в Excel: {filepath}")
    return filepath
